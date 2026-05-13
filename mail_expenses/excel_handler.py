import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import config


# Заголовки столбцов — точно как в оригинальной таблице
COLUMN_HEADERS = [
    '№',
    'Дата оплаты/отправки',
    'Номер платежного документа / РПО',
    'Назначение расходов',
    'Отправитель/Плательщик',
    'Получатель',
    'Сумма',
    'Наименование Проекта',
]

# Ширина столбцов (символов)
COLUMN_WIDTHS = [6, 18, 28, 18, 22, 35, 12, 25]


def export_to_excel(records, filter_params=None):
    """
    Экспортирует записи в Excel-файл.
    Структура полностью соответствует оригинальной таблице «Расходы по проектам».
    Возвращает путь к созданному файлу.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Общие расходы'

    _write_header(ws)
    _write_data(ws, records)
    _apply_column_widths(ws)

    # Закрепить строки с шапкой
    ws.freeze_panes = 'A3'

    export_dir = os.path.join(os.path.dirname(config.DATABASE_PATH), 'exports')
    os.makedirs(export_dir, exist_ok=True)

    now = datetime.now()
    suffix = _build_filename_suffix(filter_params)
    filename = f'Расходы_по_проектам_{suffix}_{now.strftime("%Y%m%d_%H%M%S")}.xlsx'
    export_path = os.path.join(export_dir, filename)
    wb.save(export_path)

    return export_path


def _write_header(ws):
    """
    Строка 1: объединённый заголовок «Расходы по проектам».
    Строка 2: названия столбцов.
    """
    num_cols = len(COLUMN_HEADERS)
    last_col = get_column_letter(num_cols)

    # Строка 1 — название таблицы
    ws.merge_cells(f'A1:{last_col}1')
    title_cell = ws['A1']
    title_cell.value = 'Расходы по проектам'
    title_cell.font = Font(name='Times New Roman', bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Стили для строки заголовков столбцов
    header_font = Font(name='Times New Roman', bold=True, size=11)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center',
                             wrap_text=True)
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Строка 2 — заголовки столбцов
    for col_idx, header in enumerate(COLUMN_HEADERS, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    ws.row_dimensions[2].height = 40


def _write_data(ws, records):
    """Заполняет строки данными начиная с 3-й строки."""
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    normal_font = Font(name='Times New Roman', size=11)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Чередование цвета строк (как в оригинале)
    even_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    for row_offset, record in enumerate(records):
        excel_row = row_offset + 3
        is_even = (row_offset % 2 == 1)

        def cell(col, value, align=None, number_format=None):
            c = ws.cell(row=excel_row, column=col, value=value)
            c.font = normal_font
            c.border = border
            c.alignment = align or left
            if is_even:
                c.fill = even_fill
            if number_format:
                c.number_format = number_format
            return c

        # A: порядковый номер (из поля seq_num, сохранённого при добавлении записи)
        cell(1, record.get('seq_num') or '', align=center)

        # B: дата
        cell(2, record.get('date_send', ''), align=center)

        # C: РПО — сохраняем как текст, чтобы Excel не трогал длинные числа
        rpo_cell = cell(3, str(record.get('rpo', '')), align=center)
        rpo_cell.number_format = '@'

        # D: назначение расходов — всегда «отправка почты»
        cell(4, 'отправка почты', align=center)

        # E: кто отправил
        cell(5, record.get('sender', '') or '', align=left)

        # F: получатель
        cell(6, record.get('recipient', '') or '', align=left)

        # G: сумма
        amount = record.get('amount', 0) or 0
        amount_cell = cell(7, float(amount), align=center, number_format='#,##0.00')

        # H: проект
        cell(8, record.get('project', '') or '', align=left)

        ws.row_dimensions[excel_row].height = 18

    # Строка итогов (сумма) после данных
    if records:
        total_row = len(records) + 3
        thin = Side(style='thin', color='000000')
        total_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        bold_font = Font(name='Times New Roman', bold=True, size=11)
        total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        for col in range(1, 9):
            c = ws.cell(row=total_row, column=col, value='')
            c.border = total_border
            c.fill = total_fill
            c.font = bold_font

        ws.cell(row=total_row, column=5).value = 'ИТОГО:'
        ws.cell(row=total_row, column=5).alignment = Alignment(horizontal='right', vertical='center')

        total_amount = sum(float(r.get('amount', 0) or 0) for r in records)
        total_cell = ws.cell(row=total_row, column=7, value=total_amount)
        total_cell.number_format = '#,##0.00'
        total_cell.alignment = Alignment(horizontal='center', vertical='center')

        count_cell = ws.cell(row=total_row, column=8,
                             value=f'Записей: {len(records)}')
        count_cell.alignment = Alignment(horizontal='left', vertical='center')

        ws.row_dimensions[total_row].height = 22


def _apply_column_widths(ws):
    for col_idx, width in enumerate(COLUMN_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def append_to_master_excel(record):
    """
    Добавляет одну строку с новой записью в основной Excel-файл на сервере.
    Файл должен существовать и иметь лист с названием из config.EXCEL_MASTER_SHEET.

    Возвращает порядковый номер добавленной строки.
    Выбрасывает RuntimeError если файл недоступен или занят.
    """
    master_file = getattr(config, 'EXCEL_MASTER_FILE', '')
    if not master_file:
        return None  # Запись в Excel отключена в настройках

    if not os.path.exists(master_file):
        raise RuntimeError(
            f'Основной Excel-файл не найден: {master_file}\n'
            'Проверьте путь EXCEL_MASTER_FILE в config.py и доступность диска R:'
        )

    try:
        wb = openpyxl.load_workbook(master_file)
    except PermissionError:
        raise RuntimeError(
            'Файл Excel открыт другим пользователем. '
            'Закройте файл и попробуйте снова.'
        )
    except Exception as e:
        raise RuntimeError(f'Не удалось открыть файл Excel: {e}')

    # Выбираем нужный лист
    sheet_name = getattr(config, 'EXCEL_MASTER_SHEET', 'Общие расходы')
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        # Берём первый лист и предупреждаем
        ws = wb.active

    # Ищем последнюю строку с данными (пропускаем заголовки в строках 1-2)
    last_data_row = 2
    for row_idx in range(ws.max_row, 2, -1):
        # Считаем строку заполненной, если есть дата (колонка B)
        if ws.cell(row=row_idx, column=2).value is not None:
            last_data_row = row_idx
            break

    new_row = last_data_row + 1

    # Следующий порядковый номер = последний № + 1
    last_seq = ws.cell(row=last_data_row, column=1).value
    try:
        seq_num = int(last_seq) + 1
    except (TypeError, ValueError):
        seq_num = record.get('seq_num') or 1

    # Стили — максимально близко к оригинальной таблице
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    font = Font(name='Times New Roman', size=11)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def _c(col, value, align=None, num_format=None):
        c = ws.cell(row=new_row, column=col, value=value)
        c.font = font
        c.border = border
        c.alignment = align or left
        if num_format:
            c.number_format = num_format
        return c

    # A: порядковый №
    _c(1, seq_num, align=center)
    # B: дата отправки
    _c(2, record.get('date_send', ''), align=center)
    # C: РПО — как текст, чтобы Excel не обрезал цифры
    rpo_cell = _c(3, str(record.get('rpo', '')), align=center)
    rpo_cell.number_format = '@'
    # D: назначение расходов (всегда одинаковое)
    _c(4, 'отправка почты', align=center)
    # E: кто отправил
    _c(5, record.get('sender', '') or '', align=left)
    # F: кому отправил
    _c(6, record.get('recipient', '') or '', align=left)
    # G: сумма
    _c(7, float(record.get('amount', 0) or 0), align=center, num_format='#,##0.00')
    # H: проект
    _c(8, record.get('project', '') or '', align=left)

    ws.row_dimensions[new_row].height = 18

    try:
        wb.save(master_file)
    except PermissionError:
        raise RuntimeError(
            'Не удалось сохранить файл Excel — он открыт другим пользователем. '
            'Запись в базу данных сохранена, но в Excel строка не добавлена. '
            'Закройте Excel-файл и повторите попытку.'
        )

    return seq_num


def _build_filename_suffix(filter_params):
    """Формирует суффикс имени файла по параметрам фильтра."""
    if not filter_params:
        return 'все'

    parts = []
    year = filter_params.get('year')
    quarter = filter_params.get('quarter')
    month = filter_params.get('month')
    project = filter_params.get('project')

    if year:
        parts.append(str(year))
    if quarter:
        parts.append(f'Q{quarter}')
    elif month:
        parts.append(f'M{month:02d}' if isinstance(month, int) else f'M{month}')
    if project:
        # Убираем недопустимые символы для имени файла
        safe = ''.join(c for c in project if c.isalnum() or c in ' _-')[:20].strip()
        if safe:
            parts.append(safe)

    return '_'.join(parts) if parts else 'все'
